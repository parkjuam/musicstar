import streamlit as st
import pandas as pd
from streamlit_drawable_canvas import st_canvas
from PIL import Image
import os
import io
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment


# --- 여기에 파일 위치를 미리 지정합니다 ---
과목_행_번호 = 12
이름_열_이름 = '성명'
# -----------------------------------------

# --- 기본 화면 설정 ---
st.set_page_config(layout="wide")
st.title("✅ 최종 서명 확인 및 엑셀 저장 앱")

# --- 서명 저장 폴더 생성 ---
if not os.path.exists("signatures"):
    os.makedirs("signatures")

# --- 앱이 재실행되어도 데이터를 기억하기 위한 공간 (세션) ---
if 'df' not in st.session_state:
    st.session_state.df = None
if 'signatures' not in st.session_state:
    st.session_state.signatures = {}
if 'original_file_bytes' not in st.session_state:
    st.session_state.original_file_bytes = None
if 'final_excel' in st.session_state:
    del st.session_state.final_excel

# --- 1. 파일 업로드 ---
uploaded_file = st.file_uploader("📂 성적 엑셀 파일을 업로드하세요.", type=["xlsx"])

if uploaded_file:
    # 새 파일이 올라오면 모든 데이터 초기화
    if st.session_state.original_file_bytes is None or uploaded_file.getvalue() != st.session_state.original_file_bytes:
        st.session_state.original_file_bytes = uploaded_file.getvalue()
        df = pd.read_excel(io.BytesIO(st.session_state.original_file_bytes), header=과목_행_번호 - 1)
        df_cleaned = df.loc[:, ~df.columns.str.contains('^Unnamed')]
        df_cleaned.dropna(subset=[이름_열_이름], inplace=True)
        st.session_state.df = df_cleaned
        st.session_state.signatures = {}
        st.rerun()

# --- 데이터가 로드된 경우에만 나머지 화면 표시 ---
if st.session_state.df is not None:
    df = st.session_state.df
    student_names = df[이름_열_이름].unique().tolist()
    
    col1, col2 = st.columns([1, 1])

    with col1:
        st.subheader("1단계: 학생 선택 및 서명 ✍️")
        search_term = st.selectbox("서명할 학생을 선택하세요:", options=student_names, index=None, placeholder="학생 선택...")

        if search_term:
            st.write(f"**'{search_term}'** 학생의 성적입니다.")
            st.dataframe(df[df[이름_열_이름] == search_term])

            # --- ✨ 여기가 수정되었습니다 (배경을 투명하게) ✨ ---
            canvas_result = st_canvas(
                stroke_width=3,
                stroke_color="#000000",
                background_color="#FFFFFF",  # 서명할 때는 다시 흰색 배경으로 변경
                height=150,
                width=400,
                drawing_mode="freedraw",
                key=f"canvas_{search_term}"
            )

            if st.button(f"'{search_term}' 학생 서명 완료하기"):
                if canvas_result.image_data is not None:
                    # --- ✨ 여기가 수정되었습니다 (이미지 배경 투명 처리 로직) ✨ ---
                    # 1. 서명 데이터를 PIL 이미지 객체로 변환
                    img_data = canvas_result.image_data
                    pil_img = Image.fromarray(img_data.astype('uint8'), 'RGBA')

                    # 2. 이미지의 흰색 배경을 투명하게 변경
                    new_data = []
                    for item in pil_img.getdata():
                        # 흰색 픽셀(R,G,B가 255)을 만나면 투명하게 (Alpha=0) 변경
                        if item[0] == 255 and item[1] == 255 and item[2] == 255:
                            new_data.append((255, 255, 255, 0))
                        else:
                            new_data.append(item)
                    
                    pil_img.putdata(new_data)

                    # 3. 투명해진 이미지를 파일로 저장
                    file_name = f"{search_term}_sign.png"
                    save_path = os.path.join("signatures", file_name)
                    pil_img.save(save_path)
                    
                    st.session_state.signatures[search_term] = save_path
                    st.success(f"'{search_term}' 학생 서명이 추가되었습니다!")
                    st.rerun() # 서명 후 UI가 바로 업데이트 되도록 추가
                else:
                    st.warning("서명을 먼저 해주세요.")

    with col2:
        st.subheader("2단계: 서명 현황 확인 ✅")
        if not st.session_state.signatures:
            st.info("아직 서명한 학생이 없습니다.")
        else:
            st.write(f"**총 {len(st.session_state.signatures)}명 서명 완료**")
            for name, path in st.session_state.signatures.items():
                st.image(path, caption=name, width=150)
    
    st.markdown("---")
    st.subheader("3단계: 전체 서명 완료 및 엑셀 저장 💾")

    if st.button("**(중요) 모든 서명을 엑셀 파일 하나로 합치기**", type="primary"):
        if not st.session_state.signatures:
            st.error("저장할 서명이 없습니다. 먼저 학생 서명을 완료해주세요.")
        else:
            with st.spinner('엑셀 파일에 서명을 삽입하는 중입니다... 잠시만 기다려주세요.'):
                try:
                    wb = load_workbook(io.BytesIO(st.session_state.original_file_bytes))
                    ws = wb.active

                    header_df = pd.read_excel(io.BytesIO(st.session_state.original_file_bytes), header=과목_행_번호 - 1)
                    if '비고' in header_df.columns:
                        remarks_col_idx = header_df.columns.tolist().index('비고') + 1
                    else:
                        remarks_col_idx = ws.max_column + 1
                        ws.cell(row=과목_행_번호, column=remarks_col_idx, value='비고')

                    for name, img_path in st.session_state.signatures.items():
                        student_row_indices = df[df[이름_열_이름] == name].index
                        if len(student_row_indices) > 0:
                            student_row_idx = student_row_indices[0] + 과목_행_번호 + 1
                            
                            img = OpenpyxlImage(img_path)
                            
                            img.height = 19 
                            img.width = 35
                            
                            cell_address = f"{get_column_letter(remarks_col_idx)}{student_row_idx}"
                            ws.add_image(img, cell_address)
                            
                            ws.row_dimensions[student_row_idx].height = 15

                            cell = ws[cell_address]
                            cell.alignment = Alignment(horizontal='center', vertical='center')

                    output_buffer = io.BytesIO()
                    wb.save(output_buffer)
                    st.session_state.final_excel = output_buffer.getvalue()
                    st.success("엑셀 파일 생성 완료! 아래 버튼으로 다운로드하세요.")
                except Exception as e:
                    st.error(f"엑셀 파일 생성 중 오류 발생: {e}")

    if 'final_excel' in st.session_state:
        st.download_button(
            label="📄 서명이 포함된 최종 엑셀 파일 다운로드",
            data=st.session_state.final_excel,
            file_name="서명완료_성적표.xlsx"
        )